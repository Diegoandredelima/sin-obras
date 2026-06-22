"""
SIN-Obras — Serviço de Exportação de Relatórios

Gera arquivos XLSX e PDF com dados do relatório-resumo.
"""
from datetime import date
from io import BytesIO

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.cadastro import Empresa
from app.models.obra import Contrato, Obra

STATUS_LABELS = {
    "EM_EXECUCAO": "Em Execução",
    "PARALISADA": "Paralisada",
    "CONCLUIDA": "Concluída",
    "PLANEJADA": "Planejada",
}

HEADER_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1B5E20")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


async def _fetch_data(db: AsyncSession) -> dict:
    total_obras = await db.scalar(select(func.count(Obra.id)).where(Obra.ativo == True))
    total_contratos = await db.scalar(select(func.count(Contrato.id)))
    total_empresas = await db.scalar(select(func.count(Empresa.id)))
    valor_total = await db.scalar(select(func.coalesce(func.sum(Contrato.valor_global), 0)))

    status_rows = await db.execute(
        select(Obra.status, func.count(Obra.id))
        .where(Obra.ativo == True, Obra.status.is_not(None))
        .group_by(Obra.status)
    )
    por_status = [
        (STATUS_LABELS.get(row[0], row[0]), row[1])
        for row in status_rows
    ]

    orgao_rows = await db.execute(
        select(Obra.orgao, func.count(Obra.id), func.coalesce(func.sum(Contrato.valor_global), 0))
        .join(Contrato, Obra.contrato_id == Contrato.id, isouter=True)
        .where(Obra.ativo == True)
        .group_by(Obra.orgao)
        .order_by(func.count(Obra.id).desc())
        .limit(15)
    )
    por_orgao = [
        (row[0] or "Não informado", row[1], float(row[2]))
        for row in orgao_rows
    ]

    return {
        "total_obras": total_obras or 0,
        "total_contratos": total_contratos or 0,
        "total_empresas": total_empresas or 0,
        "valor_total": float(valor_total or 0),
        "por_status": por_status,
        "por_orgao": por_orgao,
    }


def gerar_xlsx(data: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório SIN-Obras"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"Relatório de Obras — SIN-Obras | Gerado em {date.today().isoformat()}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A3:E3")
    ws["A3"] = "Indicadores Gerais"
    ws["A3"].font = Font(bold=True, size=12)
    ws.row_dimensions[3].height = 22

    indicadores = [
        ("Total de Obras", data["total_obras"]),
        ("Total de Contratos", data["total_contratos"]),
        ("Total de Empresas", data["total_empresas"]),
        ("Valor Total dos Contratos", f"R$ {data['valor_total']:,.2f}"),
    ]
    for i, (label, val) in enumerate(indicadores, 4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=11)
        ws.cell(row=i, column=3, value=val).font = Font(size=11)
        ws.cell(row=i, column=1).border = THIN_BORDER
        ws.cell(row=i, column=3).border = THIN_BORDER

    row = 9
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1, value="Obras por Status").font = Font(bold=True, size=12)
    row += 1
    for col, h in enumerate(["Status", "Total"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")
    row += 1
    for status, total in sorted(data["por_status"], key=lambda x: x[1], reverse=True):
        ws.cell(row=row, column=1, value=status).border = THIN_BORDER
        ws.cell(row=row, column=2, value=total).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        row += 1

    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="Top 15 Órgãos por Volume de Obras").font = Font(bold=True, size=12)
    row += 1
    for col, h in enumerate(["Órgão", "Total Obras", "Valor Total (R$)"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")
    row += 1
    for orgao, tot, val in data["por_orgao"]:
        ws.cell(row=row, column=1, value=orgao).border = THIN_BORDER
        ws.cell(row=row, column=2, value=tot).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=f"R$ {val:,.2f}").border = THIN_BORDER
        row += 1

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


SITUACAO_LABELS_EXPORT = {
    "EM_ANDAMENTO": "Em Andamento", "CONCLUIDA": "Concluída",
    "PARALISADA": "Paralisada", "A_INICIAR": "A Iniciar",
    "INACABADA": "Inacabada", "RESCINDIDA": "Rescindida",
    "ARQUIVADA": "Arquivada",
}


def gerar_xlsx_obras(obras: list) -> BytesIO:
    """Gera planilha XLSX a partir da lista de obras filtradas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Obras"

    headers = [
        "Título", "Município", "Órgão", "Empresa", "Nº Contrato",
        "Situação", "Status", "% Execução",
        "Valor Contrato (R$)", "Valor Final (R$)", "Valor Medido (R$)",
        "Vigência Início", "Vigência Fim",
    ]
    n_cols = len(headers)

    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    ws["A1"] = f"Lista de Obras — SIN-Obras | Gerado em {date.today().isoformat()}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    for row_idx, o in enumerate(obras, 3):
        row_data = [
            o.titulo,
            o.municipio or "",
            o.orgao or "",
            o.empresa_razao_social or "",
            o.contrato_numero or "",
            SITUACAO_LABELS_EXPORT.get(o.situacao or "", o.situacao or ""),
            STATUS_LABELS.get(o.status or "", o.status or ""),
            float(o.percentual_executado or 0),
            float(o.valor_contrato or 0),
            float(o.valor_final or o.valor_global or o.valor_contrato or 0),
            float(o.valor_medido or 0),
            str(o.vigencia_inicio or ""),
            str(o.vigencia_fim or ""),
        ]
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = THIN_BORDER
            if col_idx in (9, 10, 11):
                c.number_format = "#,##0.00"

    for col_idx, w in enumerate([50, 18, 15, 40, 18, 18, 15, 12, 20, 20, 20, 15, 15], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def gerar_pdf(data: dict) -> BytesIO:
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_font("DejaVu", "", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", uni=True)
    pdf.add_font("DejaVu", "B", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", uni=True)
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.set_font("DejaVu", "B", 16)
    pdf.set_text_color(27, 94, 32)
    pdf.cell(0, 10, "Relatorio de Obras — SIN-Obras", ln=True, align="C")
    pdf.set_font("DejaVu", "", 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 8, f"Gerado em {date.today().isoformat()}", ln=True, align="C")
    pdf.ln(5)

    pdf.set_font("DejaVu", "B", 13)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 10, "Indicadores Gerais", ln=True)
    pdf.set_font("DejaVu", "", 11)
    indicadores = [
        ("Total de Obras:", str(data["total_obras"])),
        ("Total de Contratos:", str(data["total_contratos"])),
        ("Total de Empresas:", str(data["total_empresas"])),
        ("Valor Total dos Contratos:", f"R$ {data['valor_total']:,.2f}"),
    ]
    for label, val in indicadores:
        pdf.cell(80, 8, label)
        pdf.cell(0, 8, val, ln=True)
    pdf.ln(5)

    pdf.set_font("DejaVu", "B", 13)
    pdf.cell(0, 10, "Obras por Status", ln=True)
    pdf.set_font("DejaVu", "B", 10)
    pdf.set_fill_color(27, 94, 32)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(100, 8, "Status", border=1, fill=True)
    pdf.cell(30, 8, "Total", border=1, fill=True, align="C")
    pdf.ln()
    pdf.set_font("DejaVu", "", 10)
    pdf.set_text_color(0, 0, 0)
    for status, total in sorted(data["por_status"], key=lambda x: x[1], reverse=True):
        pdf.cell(100, 7, status, border=1)
        pdf.cell(30, 7, str(total), border=1, align="C")
        pdf.ln()
    pdf.ln(8)

    pdf.set_font("DejaVu", "B", 13)
    pdf.cell(0, 10, "Top 15 Orgaos por Volume de Obras", ln=True)
    pdf.set_font("DejaVu", "B", 10)
    pdf.set_fill_color(27, 94, 32)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(100, 8, "Orgao", border=1, fill=True)
    pdf.cell(30, 8, "Obras", border=1, fill=True, align="C")
    pdf.cell(40, 8, "Valor (R$)", border=1, fill=True, align="C")
    pdf.ln()
    pdf.set_font("DejaVu", "", 9)
    pdf.set_text_color(0, 0, 0)
    for orgao, tot, val in data["por_orgao"]:
        pdf.cell(100, 7, orgao[:50], border=1)
        pdf.cell(30, 7, str(tot), border=1, align="C")
        pdf.cell(40, 7, f"R$ {val:,.0f}", border=1, align="R")
        pdf.ln()

    output = BytesIO()
    pdf.output(output)
    output.seek(0)
    return output
