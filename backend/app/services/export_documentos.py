"""
SIN-Obras — Exportação de Documentos de Objeto (XLS)

Gera os arquivos XLSX dos documentos de objeto (Boletim de Medição, Memória de
Cálculo e RDO) reaproveitando o cálculo de boletim de ``services.portal`` e o
padrão visual de ``services.export_relatorio`` (cabeçalho azul SIN, bordas).
Os PDFs equivalentes são gerados no frontend pelas rotas de impressão
(PrintLayout), portanto aqui só tratamos do formato XLS.
"""
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.objeto import Contrato, Objeto
from app.models.portal import DiarioObra, MedicaoItem
from app.services import portal as portal_service

HEADER_FILL = PatternFill(start_color="1B3C73", end_color="1B3C73", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=13, color="1B3C73")
LABEL_FONT = Font(name="Calibri", bold=True, size=10, color="555555")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

TEMPO_LABEL = {"BOM": "Bom", "CHUVA_FRACA": "Chuva fraca", "CHUVA_FORTE": "Chuva forte"}


def _save(wb: Workbook) -> BytesIO:
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _header_inst(ws, ncols: int, titulo: str, objeto_titulo: str, extra: list[str]) -> int:
    """Escreve o cabeçalho institucional e devolve a próxima linha livre."""
    last = get_column_letter(ncols)
    linhas = [
        "GOVERNO DO ESTADO DO RIO GRANDE DO NORTE",
        "SECRETARIA DE ESTADO DA INFRAESTRUTURA — SIN",
        titulo,
    ]
    row = 1
    for i, texto in enumerate(linhas):
        ws.merge_cells(f"A{row}:{last}{row}")
        cell = ws.cell(row=row, column=1, value=texto)
        cell.alignment = CENTER
        cell.font = TITLE_FONT if i == 2 else Font(name="Calibri", bold=True, size=10)
        row += 1
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws.cell(row=row, column=1, value=f"Objeto: {objeto_titulo}")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.font = Font(name="Calibri", size=10)
    row += 1
    for texto in extra:
        ws.merge_cells(f"A{row}:{last}{row}")
        c = ws.cell(row=row, column=1, value=texto)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.font = Font(name="Calibri", size=10)
        row += 1
    return row + 1


# ---------------------------------------------------------------------------
# Boletim de Medição
# ---------------------------------------------------------------------------
async def _ctx_objeto(db: AsyncSession, objeto_id: UUID, contrato_id: UUID | None) -> tuple[str, str | None, str | None]:
    objeto = await db.scalar(select(Objeto).where(Objeto.id == objeto_id))
    contrato = await db.scalar(select(Contrato).where(Contrato.id == contrato_id)) if contrato_id else None
    return (
        objeto.titulo if objeto else "—",
        contrato.numero_contrato if contrato else None,
        contrato.objeto if contrato else None,
    )


async def gerar_boletim_xlsx(db: AsyncSession, medicao_id: UUID) -> tuple[BytesIO, str]:
    medicao = await portal_service.get_medicao_by_id(db, medicao_id)
    boletim = await portal_service.montar_boletim(db, medicao)
    objeto_titulo, contrato_num, _ = await _ctx_objeto(db, medicao.objeto_id, medicao.contrato_id)

    wb = Workbook()
    ws = wb.active
    ws.title = f"BM {medicao.numero_medicao:02d}"

    cols = [
        ("ITEM", 8), ("DESCRIÇÃO", 50), ("UNID.", 8),
        ("QTD PREV.", 12), ("QTD PERÍODO", 12), ("QTD ACUM.", 12), ("QTD SALDO", 12),
        ("% PER.", 9), ("% ACUM.", 9), ("PREÇO UNIT.", 14),
        ("VLR PERÍODO", 14), ("VLR ACUM.", 14), ("SALDO", 14),
    ]
    numero_art = boletim.get("numero_art")
    extra = [
        f"Contrato: {contrato_num or '—'}    ART/RRT: {numero_art or '—'}"
        f"    Boletim da {medicao.numero_medicao}ª Medição"
    ]
    row = _header_inst(ws, len(cols), "BOLETIM DE MEDIÇÃO", objeto_titulo, extra)

    for i, (titulo, larg) in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=titulo)
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, THIN_BORDER
        ws.column_dimensions[get_column_letter(i)].width = larg
    row += 1

    for idx, it in enumerate(boletim["itens"], start=1):
        valores = [
            idx, it["descricao"] or "", it["unidade"] or "",
            float(it["quantidade_prevista"]), float(it["quantidade_periodo"]),
            float(it["quantidade_acumulada"]), float(it["quantidade_saldo"]),
            float(it["percentual_periodo"]) / 100, float(it["percentual_acumulado"]) / 100,
            float(it["valor_unitario"]), float(it["valor_bruto"]),
            float(it["acumulado_atual"]), float(it["saldo"]),
        ]
        for c, val in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = THIN_BORDER
            if c == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif c >= 4:
                cell.alignment = RIGHT
                cell.number_format = "0.00%" if c in (8, 9) else ("#,##0.00" if c >= 10 else "#,##0.0000")
            else:
                cell.alignment = CENTER
        row += 1

    # Totais
    row += 1
    def _total(label: str, valor) -> None:
        nonlocal row
        ws.merge_cells(f"A{row}:J{row}")
        lc = ws.cell(row=row, column=1, value=label)
        lc.alignment, lc.font = RIGHT, LABEL_FONT
        vc = ws.cell(row=row, column=11, value=float(valor))
        ws.merge_cells(f"K{row}:M{row}")
        vc.alignment, vc.number_format, vc.font = RIGHT, "#,##0.00", Font(name="Calibri", bold=True)
        row += 1

    _total("Valor bruto total", boletim["valor_bruto_total"])
    _total("Faturamento direto", boletim["valor_faturamento_direto"])
    _total(f"Retenção ({boletim['percentual_retencao']}%)", boletim["retencao"])
    _total("Valor líquido", boletim["valor_liquido"])

    return _save(wb), f"boletim_medicao_{medicao.numero_medicao:02d}.xlsx"


# ---------------------------------------------------------------------------
# Memória de Cálculo
# ---------------------------------------------------------------------------
async def gerar_memoria_xlsx(db: AsyncSession, medicao_id: UUID) -> tuple[BytesIO, str]:
    medicao = await portal_service.get_medicao_by_id(db, medicao_id)
    # Carrega itens com memória e evento.
    itens = (await db.scalars(
        select(MedicaoItem)
        .where(MedicaoItem.medicao_id == medicao_id)
        .options(selectinload(MedicaoItem.memoria), selectinload(MedicaoItem.evento))
    )).all()
    objeto_titulo, contrato_num, _ = await _ctx_objeto(db, medicao.objeto_id, medicao.contrato_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "M. DE CÁLCULO"

    cols = [("ITEM", 10), ("DESCRIÇÃO", 46), ("UNID.", 8), ("C/P", 10), ("L", 10),
            ("H", 10), ("N", 8), ("%", 8), ("QTD (A/V)", 14)]
    extra = [
        f"Contrato: {contrato_num or '—'}    Memória de Cálculo — {medicao.numero_medicao}ª Medição",
        "LEGENDA: C/P = COMPRIMENTO/PERÍMETRO; L = LARGURA; H = ALTURA; N = Nº DE REPETIÇÕES; A/V = ÁREA/VOLUME",
    ]
    row = _header_inst(ws, len(cols), "MEMÓRIA DE CÁLCULO", objeto_titulo, extra)

    for i, (titulo, larg) in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=titulo)
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, THIN_BORDER
        ws.column_dimensions[get_column_letter(i)].width = larg
    row += 1

    for idx, item in enumerate(itens, start=1):
        ev = item.evento
        # Linha de cabeçalho do item.
        ws.cell(row=row, column=1, value=str(idx)).border = THIN_BORDER
        dc = ws.cell(row=row, column=2, value=ev.descricao if ev else "")
        dc.font = Font(name="Calibri", bold=True)
        for c in range(1, len(cols) + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=3, value=ev.unidade if ev else "")
        row += 1
        for linha in item.memoria:
            vals = [
                "", linha.descricao or "", "",
                _f(linha.comprimento), _f(linha.largura), _f(linha.altura),
                _f(linha.n_repeticoes), _f(linha.percentual), float(linha.quantidade),
            ]
            for c, val in enumerate(vals, start=1):
                cell = ws.cell(row=row, column=c, value=val)
                cell.border = THIN_BORDER
                if c >= 4:
                    cell.alignment, cell.number_format = RIGHT, "#,##0.0000"
                elif c == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            row += 1

    return _save(wb), f"memoria_calculo_{medicao.numero_medicao:02d}.xlsx"


def _f(value) -> float | None:
    return float(value) if value is not None else None


# ---------------------------------------------------------------------------
# RDO — Relatório Diário de Objeto
# ---------------------------------------------------------------------------
async def gerar_rdo_xlsx(db: AsyncSession, diario_id: UUID) -> tuple[BytesIO, str]:
    diario = await db.scalar(select(DiarioObra).where(DiarioObra.id == diario_id))
    if diario is None:
        from fastapi import HTTPException, status
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Registro de diário não encontrado.")
    objeto_titulo, contrato_num, _ = await _ctx_objeto(db, diario.objeto_id, None)

    wb = Workbook()
    ws = wb.active
    ws.title = "RDO"
    for col, larg in (("A", 28), ("B", 16), ("C", 28), ("D", 16)):
        ws.column_dimensions[col].width = larg

    extra = [
        f"Data: {diario.data_registro.strftime('%d/%m/%Y')}    Contrato: {contrato_num or '—'}",
        f"Tempo — Manhã: {TEMPO_LABEL.get(diario.tempo_manha or '', '—')}   "
        f"Tarde: {TEMPO_LABEL.get(diario.tempo_tarde or '', '—')}   "
        f"Pluviometria: {diario.pluviometria_mm or 0} mm",
    ]
    row = _header_inst(ws, 4, "RELATÓRIO DIÁRIO DE OBRA (RDO)", objeto_titulo, extra)

    def _tabela(titulo_a: str, titulo_b: str, linhas: list[dict], campo: str) -> None:
        nonlocal row
        for col, txt in ((1, titulo_a), (2, titulo_b)):
            cell = ws.cell(row=row, column=col, value=txt)
            cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, THIN_BORDER
        row += 1
        for linha in linhas or []:
            ws.cell(row=row, column=1, value=str(linha.get(campo, ""))).border = THIN_BORDER
            qc = ws.cell(row=row, column=2, value=linha.get("quantidade"))
            qc.alignment, qc.border = RIGHT, THIN_BORDER
            row += 1
        row += 1

    _tabela("EQUIPAMENTO", "QUANT.", diario.equipamentos_lista, "nome")
    _tabela("MÃO DE OBRA", "QUANT.", diario.mao_de_obra, "funcao")

    def _bloco(titulo: str, texto: str | None) -> None:
        nonlocal row
        c = ws.cell(row=row, column=1, value=titulo)
        c.font = LABEL_FONT
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        tc = ws.cell(row=row, column=1, value=texto or "—")
        tc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 2

    _bloco("ATIVIDADES REALIZADAS", diario.atividades_realizadas)
    _bloco("OCORRÊNCIAS", diario.ocorrencias)
    _bloco("OBSERVAÇÕES DA FISCALIZAÇÃO", diario.observacoes_fiscal)

    return _save(wb), f"rdo_{diario.data_registro.isoformat()}.xlsx"
